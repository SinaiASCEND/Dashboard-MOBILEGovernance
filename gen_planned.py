#!/usr/bin/env python3
"""Parse the MASTER Agenda Tracker (EEC Agenda + Subcommittee Agendas tabs)
into a static planned-agenda.js consumed by the mobile governance dashboard.

Usage:
    python3 gen_planned.py [path-to-tracker.xlsx]
If no path is given, it uses the most recent *Agenda_Tracker*.xlsx in the
current folder.
"""
import datetime, re, json, sys, glob, os

from openpyxl import load_workbook

if len(sys.argv) > 1:
    SRC = sys.argv[1]
else:
    cands = sorted(glob.glob("*Agenda_Tracker*.xlsx"), key=os.path.getmtime, reverse=True)
    if not cands:
        sys.exit("No tracker given and no *Agenda_Tracker*.xlsx found in this folder.")
    SRC = cands[0]
print(f"Reading: {SRC}")
wb = load_workbook(SRC, read_only=True, data_only=True)

MONTHS = {m: i for i, m in enumerate(
    ['january','february','march','april','may','june','july','august',
     'september','october','november','december'], 1)}

def clean(v):
    if v is None: return ""
    return str(v).replace("\xa0", " ").strip()

def split_title(v):
    """A title cell may pack a heading line + newline-separated sub-items
    (e.g. a policy-review item listing each policy). First non-empty line is
    the title; remaining non-empty lines become sub-items."""
    s = (str(v) if v is not None else "").replace("\xa0", " ")
    lines = [ln.strip() for ln in s.splitlines()]
    lines = [ln for ln in lines if ln]
    if not lines:
        return "", []
    return lines[0], lines[1:]

def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    s = clean(v)
    if not s: return None
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = re.search(r'([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})', s)
    if m and m.group(1).lower() in MONTHS:
        return f"{int(m.group(3)):04d}-{MONTHS[m.group(1).lower()]:02d}-{int(m.group(2)):02d}"
    return None

COMMITTEES = ("PCCS", "CCS", "CIS", "AES", "SAS", "CSS")
by = {}  # committee -> { date -> [items] }

# ---------- EEC Agenda ----------
ws = wb["EEC Agenda"]
cur = None
seq = 0
for r in ws.iter_rows(min_row=4, values_only=True):  # row 4 = header (1-indexed)
    # header row check
    if clean(r[0]) == "EEC Date":
        continue
    d = parse_date(r[0])
    if d:
        cur = d
        seq = 0
        by.setdefault("EEC", {}).setdefault(cur, [])
    title = clean(r[1])
    if not (cur and title):
        continue
    title, subitems = split_title(r[1])
    seq += 1
    item = {
        "n": seq,
        "title": title,
        "category": clean(r[2]) or None,
        "owner": clean(r[3]) or None,        # Subcommittee Owner
        "presenter": clean(r[4]) or None,
        "confirmed": clean(r[5]) or None,    # Presenter confirmed?
        "guests": clean(r[6]) or None,       # Additional guests
        "ready": clean(r[8]) if len(r) > 8 else None,  # READY TO GO?
        "planned": True,
    }
    if subitems:
        item["subitems"] = subitems
    by["EEC"][cur].append(item)

# ---------- Subcommittee Agendas ----------
ws = wb["Subcommittee Agendas"]
cursub = None
curdate = None
seq = 0
for r in ws.iter_rows(min_row=4, values_only=True):
    if clean(r[0]) == "Subcommittee":  # header
        continue
    c0 = clean(r[0])
    if c0 in COMMITTEES:
        cursub = c0
    d = parse_date(r[1])
    if d:
        curdate = d
        seq = 0
    title = clean(r[2])
    if not (cursub and curdate and title) or title == "(no items scheduled)":
        continue
    title, subitems = split_title(r[2])
    seq += 1
    item = {
        "n": seq,
        "title": title,
        "presenter": clean(r[3]) or None,
        "goesToEEC": parse_date(r[4]),
        "planned": True,
    }
    if subitems:
        item["subitems"] = subitems
    by.setdefault(cursub, {}).setdefault(curdate, []).append(item)

# Only keep committees the dashboard knows about
KEEP = ("EEC", "PCCS", "CCS", "CIS", "AES")
by = {k: by[k] for k in KEEP if k in by}

# Drop empty date buckets (meetings with no items)
for cid in list(by):
    by[cid] = {d: items for d, items in by[cid].items() if items}

payload = {
    "generated": "2026-05-18",
    "source": "MASTER Agenda Tracker (EEC Agenda + Subcommittee Agendas tabs)",
    "byCommittee": by,
}

# Summary to stdout
for cid in KEEP:
    if cid in by:
        tot = sum(len(v) for v in by[cid].values())
        print(f"{cid}: {len(by[cid])} meetings, {tot} items")

js = """// planned-agenda.js — Scheduled (planned) agenda items for upcoming governance meetings.
//
// Source: MASTER Agenda Tracker (sheet "EEC Agenda" + sheet "Subcommittee Agendas"),
// dated 2026-05-18. Generated, do not hand-edit; re-run gen_planned.py to refresh.
//
// Shape:
//   window.PLANNED_AGENDA.byCommittee[committee][YYYY-MM-DD] = [ {n, title, category,
//     owner, presenter, confirmed, guests, ready, goesToEEC, planned}, ... ]
// EEC items carry category/owner/presenter/guests/ready; subcommittee items carry
// presenter + goesToEEC (the EEC meeting this item feeds up to).

(function () {
  const DATA = %s;

  window.PLANNED_AGENDA = Object.assign({}, DATA, {
    // Planned agenda items for a given committee + date (""[]"" if none).
    itemsFor(committee, date) {
      const c = (this.byCommittee || {})[committee];
      if (!c) return [];
      return c[date] || [];
    },
    // True if this committee+date has any planned agenda on file.
    has(committee, date) {
      return this.itemsFor(committee, date).length > 0;
    },
  });
})();
""" % json.dumps(payload, ensure_ascii=False, indent=2)

with open("planned-agenda.js", "w", encoding="utf-8") as f:
    f.write(js)
print("\nWrote planned-agenda.js (%d bytes)" % len(js))
